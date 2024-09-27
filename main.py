import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
import xlwings as xw

# 輸入區域 ========================================
new_data_route = '/Users/wubaixian/Desktop/main.xlsx'
report_file = '/Users/wubaixian/Desktop/Milestones Month View 20240927.xlsx'
last_week = '0920' # ex. 0920

# 讀取資料================================================
app = xw.App(visible=False)
wb_a = xw.Book(new_data_route)
wb_b = openpyxl.load_workbook(report_file)
sheet_last_week = wb_b[last_week]
source_sheet = wb_a.sheets[0]

# 新增一個以今天日期命名的工作表 ========================================
today_str = datetime.today().strftime('%m%d')
new_sheet = wb_b.create_sheet(title=today_str)

# 複製 a 工作表的內容到新工作表 ========================================
for row in source_sheet.range('A1').expand().value:
    new_sheet.append(row)

#  移除多餘欄位 ========================================
last_week_headers = [cell.value for cell in sheet_last_week[1]]
current_week_headers = [cell.value for cell in new_sheet[1]]
extra_columns = []
for col_idx, header in enumerate(current_week_headers, start=1):
    if header not in last_week_headers:
        extra_columns.append(col_idx)

for col_idx in reversed(extra_columns):  # 從後往前刪除，避免影響列數
    new_sheet.delete_cols(col_idx)

# 插入欄位 ========================================
new_sheet.insert_cols(8)
new_sheet.insert_cols(9)
new_sheet.insert_cols(10)
new_sheet.insert_cols(11)
new_sheet.insert_cols(12)

# 複製 A1 到 AQ 欄位的資料與格式 ========================================
for row in sheet_last_week.iter_rows(min_row=1, max_row=1, min_col=1, max_col=46):
    for cell in row:
        # 複製值
        new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)

        # 複製格式
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

for col in range(1, 44):  # AQ 是第43欄
    col_letter = get_column_letter(col)
    new_sheet.column_dimensions[col_letter].width = sheet_last_week.column_dimensions[col_letter].width

# 排序 (Translated Account -> 'Opportunity ID -> Milestone ID) ========================================
data = []
for row in new_sheet.iter_rows(min_row=2, values_only=True):  # 忽略第一列的標題
    if str(row[0]).lower().startswith('https://'):
        data.append(row)

sorted_data = sorted(data, key=lambda x: (
    x[1] if x[1] is not None else '',
    x[2] if x[2] is not None else '',
    x[4] if x[4] is not None else '' 
))

for row_idx, row_data in enumerate(sorted_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        new_sheet.cell(row=row_idx, column=col_idx, value=value)

# 拉公式 ========================================
max_row = new_sheet.max_row

for row in range(2, max_row):
    new_sheet[f'H{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,4,FALSE),"New")'
    new_sheet[f'I{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,5,FALSE),"")'
    new_sheet[f'J{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,6,FALSE),"")'
    new_sheet[f'K{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,7,FALSE),"")'
    new_sheet[f'L{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,8,FALSE),"")'

wb_b._sheets.insert(0, wb_b._sheets.pop(wb_b.index(new_sheet)))
wb_b.save(report_file)

# 計算公式 ========================
wb = xw.Book(report_file)
wb.app.calculate()
wb.save(report_file)
wb.close()

# vlookup 檢查區塊 ========================================

wb_b = openpyxl.load_workbook(report_file, data_only=True)
new_sheet = wb_b[today_str]

for row in range(2, max_row):
    new_sheet[f'AP{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,4,FALSE),"New")'
    new_sheet[f'AQ{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,5,FALSE),"")'
    new_sheet[f'AR{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,6,FALSE),"")'
    new_sheet[f'AS{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,7,FALSE),"")'
    new_sheet[f'AT{row}'] = f'=IFERROR(VLOOKUP($E{row},\'{last_week}\'!$E$1:$O$500,8,FALSE),"")'

# 應用篩選功能 ========================================
new_sheet.auto_filter.ref = new_sheet.dimensions

# 設定欄位群組 ========================================
new_sheet.column_dimensions.group('A', 'A', hidden=True)
new_sheet.column_dimensions.group('C', 'C', hidden=True)
new_sheet.column_dimensions.group('E', 'E', hidden=True)
new_sheet.column_dimensions.group('G', 'G', hidden=True)
new_sheet.column_dimensions.group('M', 'N', hidden=True)
new_sheet.column_dimensions.group('P', 'P', hidden=True)
new_sheet.column_dimensions.group('R', 'T', hidden=True)

# 移除為零儲存格 ========================================
for row in range(2, max_row + 1):
    for col in ['H', 'I', 'J', 'K', 'L']:
        cell = new_sheet[f'{col}{row}']
        if cell.value == 0:
            new_sheet[f'{col}{row}'] = None  # 將儲存格設為 None 刪除其內容

# 儲存修改後的 b 檔案 ========================================
wb_b.save(report_file)
app.quit()